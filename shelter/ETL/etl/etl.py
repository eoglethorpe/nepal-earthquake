"""Module for extracting data from dropbox and aggregating"""
#TODO: better way for specifyign which folder to pull from (latest?)
##how to handle names?
##qualify sheets as being valid before cycling or skip invalid ones?
##pull from rules based text file
##log put on dbox, make better


import dropbox
import os
import cStringIO
import re
from openpyxl import load_workbook


#dropbox setup
db_key = os.environ['db_key']
db_secret = os.environ['db_secret']
db_access = os.environ['db_access']
client = dropbox.client.DropboxClient(db_access)

DB_PATH = '/2015 Nepal EQ/04 im/reporting/Database_&_Template/Incoming/7 August 2015'
TEST_FILE = '/4W worldrenew_sheltercluster_August 5, 2015 FINAL.xlsx'

def iterate_reports():
    """cycle through all reports contained in dbox directory"""

    meta = client.metadata(DB_PATH, list=True)
    file_list = [str(f['path']) for f in meta['contents'] if re.search('xls|xlsx$',str(f))]
    #file_list = [DB_PATH+"/Batas Foundation.xlsx/"]
    for f in file_list:
        #pull down workbook from specified directory
        print "pulling! " + f
        wb_current = pull_wb(f)

        #clean the workboook
        print "cleaning " + f
        clean_file(wb_current, f)


def clean_file(wb, path):
    #TODO: log
    """cycle through reports and apply cleaning algorithms"""
    
    #get our two sheets
    db = wb.get_sheet_by_name('Database')
    ref = wb.get_sheet_by_name('Reference')

    #setup log

    #####do edit stuff


    #***Column A must be in Reference>ImplementingAgency 
    ##if not: make a note
    db_col_loc = find_in_header(db, 'Implementing agency')
    ref_col_loc = find_in_header(ref,'Implementing_Agency_Name')

    missing_names = colvals_notincol(db, db_col_loc, ref, ref_col_loc)

    #upload with name of file at end
    #client.put_file(DB_PATH + '/edited/' + path.rsplit('/', 1)[1], pull_file(path))
    print 'uploaded! ' + path

new_report = False
def report_a_log(log_value, path):
    """write out contents for a given log - ends if a new path is given"""

    #if module is starting and we haven't logged anything
    if 'current_path' not in locals():
        current_path = path.rsplit('/', 1)[1]
        new_report = True
    
    #if we are recieving a new path
    elif current_path != path.rsplit('/', 1)[1]:
        current_path = path.rsplit('/', 1)[1] 
        new_report = True

    if new_report:
        #write out
        with open('/Users/ewanog/code/nepal-earthquake/shelter/etl/etl/' 
            + path.rsplit('/', 1)[1] + '.txt', "w") as f:
            for log in current_log:
                

        current_log = [log_value]





def find_in_header(sheet, find_val):
    """find the coordinate of a value in header (assumes header is in row 1)"""
    for row in sheet.iter_rows('A1:' + find_last_value(sheet,'A','r')):
        for cell in row:
            if cell.value == find_val:
                return cell.column

    #if we haven't returned anything yet
    return None

def colvals_notincol(sheet_val,col_val,sheet_ref,col_ref):
    """return values from a column that are NOT in a reference column"""
    not_in = []
    to_search = []

    #create an array from sheet_ref with values to be searched (as opposed to nested loops)
    #iter_rows syntax: sheet.iter_rows('A1:A2')
    for row in sheet_ref.iter_rows(col_ref + "2:" + 
        find_last_value(sheet_ref, 'A', 'c')):

        for cell in row:
            to_search.append(str(cell.value))

    #now search through vals and see if they're present
    for row in sheet_val.iter_rows(col_val + "2:" + 
        find_last_value(sheet_val, 'A', 'c')):

        for cell in row:
            if cell.value not in to_search:
                not_in.append(str(cell.value))

    return not_in
             


def find_last_value(sheet, start_location, r_or_c):
    """find position of last value in a given row or column"""
    #extract form r_or_c if we should be iterating a row or column
    
    row_it = 0
    col_it = 0
    if r_or_c == 'c':
        row_it = 1
    elif r_or_c == 'r':
        col_it = 1
    else:
        raise Exception("r_or_c must be r or c!")
    
    #look for a cell without value, and if we find a blank, traverse 10 more
    #to make sure there's no blanks
    blank_count = 0
    found = False
    cur_cell = sheet[start_location+'1']
    while not found:
        cur_cell = cur_cell.offset(row = row_it, column= col_it)

        if not cur_cell.value:
            if blank_count == 0:
                #coord of a cell step back 1
                last_found = cur_cell.offset(row = -row_it, column= -col_it).coordinate
                blank_count+=1

            elif blank_count == 10:
                found = True
            else:
                blank_count+=1
        else:
            blank_count=0

    return last_found


def pull_wb(location):
    """return an excel pulled from dropbox"""

    in_mem_file = pull_file(location)

    wb = load_workbook(in_mem_file)
    print "pulled! " + str(wb.get_sheet_names())  
    in_mem_file.close()

    return wb

def pull_file(location):
    """pull a file from dropbox"""
    to_ret = cStringIO.StringIO()

    with client.get_file(location) as f:
        to_ret.write(f.read())
    f.close()

    return to_ret

def test():
    return pull_wb(DB_PATH+TEST_FILE)

if __name__ == '__main__':
    iterate_reports()
